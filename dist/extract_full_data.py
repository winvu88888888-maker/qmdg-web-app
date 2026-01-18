import openpyxl
import json

def extract_full_data(filename):
    wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    full_data = {}

    # 1. Extract Gua Attributes from 'bang_ky_mon'
    # Headers: Khái niệm, Tính tình, Hình thái, Thiên thời, Địa lý, Nhân vật, Động vật, Thực vật, Đồ ăn, Tĩnh vật, Nhân thể, tật bệnh, Thời gian, Sắc thái, xếp hạng
    sheet_bang = wb['bang_ky_mon']
    headers = [cell.value for cell in sheet_bang[1]]
    gua_data = {}
    for row in sheet_bang.iter_rows(min_row=2, values_only=True):
        if row[1]:
            name = row[1].strip()
            item = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    item[headers[i]] = str(val) if val is not None else ""
            gua_data[name] = item
    full_data['GUA_ATTRIBUTES'] = gua_data

    # 2. Extract Stem-on-Stem from 'cấu trúc trận kỳ môn'
    # This sheet has combos like "1. Mậu/Mậu", "2.Mậu/Ất", etc.
    # It seems to be organized by blocks.
    sheet_struct = wb['cấu trúc trận kỳ môn']
    stem_combos = {}
    for row in sheet_struct.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and ('. ' in cell or '.' in cell):
                # Try to parse "1. Mậu/Mậu: Phục Vịnh cách..."
                parts = cell.split(':', 1)
                header = parts[0].strip()
                content = parts[1].strip() if len(parts) > 1 else ""
                
                # Extract the "Mậu/Mậu" part
                sub_parts = header.split('.')
                if len(sub_parts) > 1:
                    combo_key = sub_parts[1].strip().replace(' ', '')
                    if '/' in combo_key:
                        stem_combos[combo_key] = content
    full_data['STEM_ON_STEM_INTERPRETATION'] = stem_combos

    # 3. Extract Door-on-Stem / Door-on-Door from 'kết hợp '
    # This sheet has a table structure.
    sheet_combine = wb['kết hợp ']
    # Let's just dump the first few rows to confirm structure
    combine_data = []
    for row in sheet_combine.iter_rows(max_row=50, values_only=True):
        combine_data.append([str(c) if c is not None else "" for c in row])
    full_data['COMBINE_TABLE_RAW'] = combine_data

    with open('qmdg_excel_full.json', 'w', encoding='utf-8') as f:
        json.dump(full_data, f, ensure_ascii=False, indent=2)
    print("Full extraction completed.")

if __name__ == "__main__":
    extract_full_data('ky_môn_độn_giap_cơ_bản123.xlsx')
