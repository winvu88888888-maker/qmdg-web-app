import openpyxl
import json

def inspect_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            result[sheet_name] = rows
        
        with open('excel_inspection.json', 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print("Inspection completed. Saved to excel_inspection.json")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_excel('ky_môn_độn_giap_cơ_bản123.xlsx')
