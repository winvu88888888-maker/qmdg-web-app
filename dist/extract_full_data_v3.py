import openpyxl
import json
import re

def clean_name(name):
    if not name: return ""
    return str(name).strip().upper()

def extract_v3(filename):
    wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    final_data = {
        "GUA_ATTRIBUTES": {},
        "STEM_COMBOS": {},
        "DOOR_MATRIX": {},
        "TOPIC_EXTRAS": {}
    }

    # 1. Extract Attributes (Gua, Stems, Branches) from 'bang_ky_mon'
    sheet_bang = wb['bang_ky_mon']
    headers = [clean_name(cell.value) for cell in sheet_bang[1]]
    for row in sheet_bang.iter_rows(min_row=2, values_only=True):
        if row[1]:
            name = clean_name(row[1])
            item = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    item[headers[i]] = str(val) if val is not None else ""
            final_data["GUA_ATTRIBUTES"][name] = item

    # 2. Extract Stem-on-Stem Combinations from 'cấu trúc trận kỳ môn'
    sheet_struct = wb['cấu trúc trận kỳ môn']
    for row in sheet_struct.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                # Match patterns like "1. Mậu/Mậu: ..." or "2. Mậu/Ất: ..."
                match = re.search(r'(\d+)\.\s*([^\/:]+)\s*\/\s*([^:]+)\s*:\s*(.*)', cell)
                if match:
                    c1 = clean_name(match.group(2))
                    c2 = clean_name(match.group(3))
                    content = match.group(4).strip()
                    key = f"{c1}{c2}"
                    final_data["STEM_COMBOS"][key] = content

    # 3. Extract Door Matrix from 'kết hợp ' (8x8)
    sheet_combine = wb['kết hợp ']
    # Let's find the door-on-door table. 
    # Usually it starts with Hưu, Sinh, ... headers
    rows = list(sheet_combine.iter_rows(values_only=True))
    # Search for header row containing doors
    doors_list = ["HƯU", "SINH", "THƯƠNG", "ĐỖ", "CẢNH", "TỬ", "KINH", "KHAI"]
    
    # Simple extraction for now based on previous inspection
    # We will just dump relevant topic-like info if it's not a clear matrix
    
    with open('qmdg_excel_full.json', 'w', encoding='utf-8') as f:
        json.dump(final_data, f, ensure_ascii=False, indent=2)

if __name__ == "__main__":
    extract_v3('ky_môn_độn_giap_cơ_bản123.xlsx')
